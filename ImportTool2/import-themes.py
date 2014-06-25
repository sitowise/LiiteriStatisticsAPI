from __future__ import print_function
from __future__ import unicode_literals

from openpyxl import load_workbook
from md5 import md5
import json
import sys
import requests

class Theme(object):

    hash = None
    parent = None
    statistics_id = None
    name = ""

class ThemeFactory(object):
    def __init__(self, repository):
        self.repository = repository
        
    def getHash(self, pieces):
        return md5('/'.join([repr(x) for x in pieces])).hexdigest()

    def getOrAdd(self, pieces, statistics_id = None):
        pieces = [x for x in pieces
            if x is not None and len(x.strip()) > 0]

        hash = self.getHash(pieces)
        results = [x for x in self.repository if x.hash == hash]
        if len(results) > 1:
            raise Exception("too many entries in repository with the same hash!")
        elif len(results) == 0:
            #print("Creating new entry")
            t = Theme()
            t.hash = hash
            t.name = pieces[-1]
            t.statistics_id = statistics_id
            self.repository.append(t)

            if len(pieces) > 1:
                parent = self.getOrAdd(pieces[:-1])
                t.parent = parent.hash

            return t
        else:
            #print("Returning existing entry")
            return results[0]

    def findByParent(self, hash):
        return (x for x in self.repository if x.parent == hash)

def restAddTheme(theme, parent):
    entry = {
        "Id": 0,
        "Name": theme.name,
        "ParentId": parent,
        }
    data = json.dumps(entry)
    url = 'http://localhost/LiiteriDataAPI/v1/themes/'
    headers = {
        'Content-type': 'application/json',
        'Accept': 'application/json',
        }
    req = requests.post(url, data, headers=headers)
    response = req.json()
    return int(response["Id"])

def restAddStatistics(theme_id, theme):
    url = 'http://localhost/LiiteriDataAPI/v1/indicators/%d' % \
        (theme.statistics_id,)
    data = None
    headers = {
        'Content-type': 'application/json',
        'Accept': 'application/json',
        }
    req = requests.get(url, headers=headers)
    obj = req.json()
    obj['ThemeId'] = theme_id

    data = json.dumps(obj)
    req = requests.put(url, data, headers=headers)
    obj = req.json()

if __name__ == '__main__':
    excel_file = r"C:\Projects\Liiteri\Liiteri_Tilastot_Rakenne.xlsx"

    wb = load_workbook(filename = excel_file, use_iterators = True)
    ws = wb.get_sheet_by_name(name = "Tilastot")

    themes = []
    tFactory = ThemeFactory(themes)

    cur_themes = {
        'Teemataso 1': '',
        'Teemataso 2': '',
        'Teemataso 3': '',
        'Teemataso 4': '',
        'Teemataso 5': '',
        }

    header = True
    headers = []
    for row in ws.iter_rows():
        if header:
            for cell in row:
                headers.append(cell.value)
            header = False
            continue

        # convert row to dictionary
        data = {}
        for i in range(len(row)):
            data[headers[i]] = row[i].value

        # header stuff
        if data["Tilasto"] == None:
            continue

        # Figure out all the themes for the current row
        for key in ['Teemataso %d' % (x+1) for x in range(5)]:
            if data[key] is None or len(data[key].strip()) == 0:
                continue
            cur_themes[key] = data[key]

            # zero pad the rest since we have a new lower level subtheme
            empty = False
            count = 0
            for tk in ['Teemataso %d' % (x+1) for x in range(5)]:
                if count > 0 and tk == key:
                    empty = True
                elif empty:
                    cur_themes[tk] = None
                count += 1

        statistics_id = data["Liiteri-Tilasto_ID"]
        if statistics_id is not None:
            statistics_id = int(statistics_id)

        theme = tFactory.getOrAdd((
                cur_themes['Teemataso 1'],
                cur_themes['Teemataso 2'],
                cur_themes['Teemataso 3'],
                cur_themes['Teemataso 4'],
                cur_themes['Teemataso 5'],
            ),
            statistics_id)

    def addThemes(parentHash = None, parentId = None, iter = 0):
        for theme in tFactory.findByParent(parentHash):
            print('%s%s' % (
                iter * ' ',
                theme.name))
            id = restAddTheme(theme, parentId)
            if theme.statistics_id is not None:
                restAddStatistics(id, theme)
            addThemes(theme.hash, id, iter + 1)
            print("addThemes");

    addThemes()

# eof